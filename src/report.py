"""
report.py - 比較結果のレポート出力モジュール（DigiKey ライフサイクル対応）

シート名・カラム名は column_config.py から動的に取得します。
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd

from src.column_config import get_column_names, get_sheet_names


# セルの背景色定義
COLOR_HEADER   = "4472C4"
COLOR_ADDED    = "C6EFCE"
COLOR_REMOVED  = "FFC7CE"
COLOR_CHANGED  = "FFEB9C"
COLOR_OBSOLETE = "FF9999"
COLOR_NRND     = "FFD580"
COLOR_ACTIVE   = "C6EFCE"
COLOR_UNKNOWN  = "D9D9D9"


def print_summary(results: dict) -> None:
    """比較結果のサマリーをコンソールに出力する"""
    cols   = get_column_names()
    qty_col = cols["quantity"]
    mfr_col = cols["manufacturer"]

    print("\n" + "=" * 50)
    print("  BOM 比較結果サマリー")
    print("=" * 50)
    print(f"\n[追加された部品]       : {len(results['added'])} 件")
    if not results["added"].empty:
        print(results["added"].to_string())
    print(f"\n[削除された部品]       : {len(results['removed'])} 件")
    if not results["removed"].empty:
        print(results["removed"].to_string())
    print(f"\n[{qty_col} 変更] : {len(results['qty_changed'])} 件")
    if not results["qty_changed"].empty:
        print(results["qty_changed"].to_string())
    print(f"\n[{mfr_col} 変更]  : {len(results['mfr_changed'])} 件")
    if not results["mfr_changed"].empty:
        print(results["mfr_changed"].to_string())
    print("\n" + "=" * 50)


def save_report(results: dict, output_path: str,
                dk_results: dict | None = None) -> None:
    """
    比較結果と DigiKey ライフサイクル情報を Excel に保存する。

    シート名は config.ini の [sheet_names] セクションから取得します。

    Args:
        results (dict): compare_bom() の戻り値
        output_path (str): 出力 Excel ファイルのパス
        dk_results (dict | None): DigiKey チェック結果
    """
    sheets = get_sheet_names()
    cols   = get_column_names()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_sheet(wb, sheets["added"],   results["added"],       COLOR_ADDED)
    _write_sheet(wb, sheets["removed"], results["removed"],     COLOR_REMOVED)
    _write_sheet(wb, sheets["qty"],     results["qty_changed"], COLOR_CHANGED)
    _write_sheet(wb, sheets["mfr"],     results["mfr_changed"], COLOR_CHANGED)

    if dk_results:
        _write_lifecycle_sheet(wb, dk_results, cols["part_number"])

    wb.save(output_path)
    print(f"\nレポートを保存しました: {output_path}")


# ============================================================
# 内部関数
# ============================================================

def _write_sheet(wb, sheet_name: str, df: pd.DataFrame,
                 fill_color: str) -> None:
    """DataFrameを1シートに書き込む"""
    ws = wb.create_sheet(title=sheet_name)

    if df.empty:
        ws["A1"] = "変更なし"
        ws["A1"].font = Font(italic=True, color="808080")
        return

    headers = [df.index.name or get_column_names()["part_number"]] + list(df.columns)
    _write_header_row(ws, headers)

    for row_idx, (part_number, row) in enumerate(df.iterrows(), start=2):
        ws.cell(row=row_idx, column=1, value=part_number)
        for col_idx, value in enumerate(row, start=2):
            ws.cell(row=row_idx, column=col_idx, value=value)
        row_fill = PatternFill("solid", start_color=fill_color)
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = row_fill

    _auto_fit_columns(ws)


def _write_lifecycle_sheet(wb, dk_results: dict,
                            part_number_col: str) -> None:
    """
    DigiKey ライフサイクル結果を2つのシートに書き込む。

    シート1「LifecycleStatus」: 全部品のステータス一覧
    シート2「Substitutes」    : 廃品・NRND部品の代替品一覧
    """
    # ── シート1: LifecycleStatus ──
    ws_lc = wb.create_sheet(title="LifecycleStatus")
    lc_headers = [part_number_col, "Status", "Lifecycle", "Error"]
    _write_header_row(ws_lc, lc_headers)

    order = {"obsolete": 0, "nrnd": 1, "active": 2, "unknown": 3}
    sorted_items = sorted(
        dk_results.items(),
        key=lambda kv: order.get(kv[1].get("lifecycle", "unknown"), 99))

    for row_idx, (pn, lc) in enumerate(sorted_items, start=2):
        lifecycle = lc.get("lifecycle", "unknown")
        ws_lc.cell(row_idx, 1, value=pn)
        ws_lc.cell(row_idx, 2, value=lc.get("status_label", "Unknown"))
        ws_lc.cell(row_idx, 3, value=lifecycle)
        ws_lc.cell(row_idx, 4, value=lc.get("error", ""))
        bg = {"obsolete": COLOR_OBSOLETE, "nrnd": COLOR_NRND,
              "active": COLOR_ACTIVE}.get(lifecycle, COLOR_UNKNOWN)
        fill = PatternFill("solid", start_color=bg)
        for c in range(1, len(lc_headers) + 1):
            ws_lc.cell(row_idx, c).fill = fill

    _auto_fit_columns(ws_lc)

    # ── シート2: Substitutes ──
    ws_sub = wb.create_sheet(title="Substitutes")
    sub_headers = [
        f"Obsolete/NRND {part_number_col}",
        "Substitute Mfr Part Number",
        "DigiKey Part Number",
        "Manufacturer Name",
        "Description",
    ]
    _write_header_row(ws_sub, sub_headers)

    row_idx = 2
    for pn, lc in sorted_items:
        lifecycle   = lc.get("lifecycle", "unknown")
        substitutes = lc.get("substitutes", [])
        if lifecycle not in ("obsolete", "nrnd"):
            continue

        bg   = COLOR_OBSOLETE if lifecycle == "obsolete" else COLOR_NRND
        fill = PatternFill("solid", start_color=bg)

        if not substitutes:
            ws_sub.cell(row_idx, 1, value=pn)
            ws_sub.cell(row_idx, 2, value="（代替品なし）")
            for c in range(1, len(sub_headers) + 1):
                ws_sub.cell(row_idx, c).fill = fill
            row_idx += 1
            continue

        for sub in substitutes:
            ws_sub.cell(row_idx, 1, value=pn)
            ws_sub.cell(row_idx, 2, value=sub.get("mfr_part_number", ""))
            ws_sub.cell(row_idx, 3, value=sub.get("digikey_part_number", ""))
            ws_sub.cell(row_idx, 4, value=sub.get("manufacturer", ""))
            ws_sub.cell(row_idx, 5, value=sub.get("description", ""))
            for c in range(1, len(sub_headers) + 1):
                ws_sub.cell(row_idx, c).fill = fill
            row_idx += 1

    if row_idx == 2:
        ws_sub["A2"] = "廃品・NRND部品なし"
        ws_sub["A2"].font = Font(italic=True, color="808080")

    _auto_fit_columns(ws_sub)


def _write_header_row(ws, headers: list) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", start_color=COLOR_HEADER)
        cell.alignment = Alignment(horizontal="center")


def _auto_fit_columns(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        col_letter = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 55)
